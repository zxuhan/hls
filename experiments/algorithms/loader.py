"""Excel workbook loader.

Reads the three sheets (Blocks, TDM, PDM) defined in ``DATA_SCHEMA.md`` and
returns an immutable list of :class:`Block` objects ready for the schedulers.

The loader is **strict**: any rule violation raises :class:`LoaderError` with
all violations collected in one pass — same behaviour as the Java backend.
"""
from __future__ import annotations

from collections import defaultdict, deque
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from .models import Block, ToolRequirement


SHEET_BLOCKS = "Blocks"
SHEET_TDM = "TDM"
SHEET_PDM = "PDM"

COL_HL_BLOCK = "HL Block"
COL_HRS = "HRS"
COL_FTE = "FTE"
COL_TOOLS = "TOOLS"
COL_TOOLS_AMOUNT = "TOOLS AMOUNT"

POSITION_AXES: tuple[str, ...] = (
    "WS",
    "(-) WS",
    "WSSC",
    "ILL",
    "RS",
    "RH Turret",
    "WH",
    "RH Library",
)

REQUIRED_BLOCKS_HEADERS: tuple[str, ...] = (
    COL_HL_BLOCK,
    COL_HRS,
    COL_FTE,
    COL_TOOLS,
    COL_TOOLS_AMOUNT,
    *POSITION_AXES,
)


@dataclass(frozen=True)
class Violation:
    sheet: str
    cell_ref: str
    code: str
    message: str

    def __str__(self) -> str:
        return f"[{self.sheet} {self.cell_ref}] {self.code}: {self.message}"


class LoaderError(RuntimeError):
    def __init__(self, violations: list[Violation]):
        self.violations = violations
        joined = "\n  ".join(str(v) for v in violations)
        super().__init__(f"{len(violations)} loader violation(s):\n  {joined}")


def _read_string(cell: Cell | None) -> str | None:
    """Return ``None`` for empty cells, else the cell's string value (untrimmed)."""
    if cell is None:
        return None
    val = cell.value
    if val is None:
        return None
    if isinstance(val, str):
        return val if val != "" else None
    return str(val)


def _read_number(cell: Cell | None) -> float | None:
    if cell is None:
        return None
    val = cell.value
    if val is None or isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    return None


def _cell_ref(row_idx_zero: int, col_idx_zero: int) -> str:
    """A1-style cell reference (1-indexed)."""
    col = ""
    n = col_idx_zero
    while True:
        col = chr(ord("A") + n % 26) + col
        n = n // 26 - 1
        if n < 0:
            break
    return f"{col}{row_idx_zero + 1}"


def _find_sheet(wb, target: str):
    """Case-insensitive, trim-aware sheet lookup."""
    needle = target.strip().lower()
    for name in wb.sheetnames:
        if name.strip().lower() == needle:
            return wb[name]
    return None


# ────────────────────────────────────────────────────────────────────────────
# Blocks sheet
# ────────────────────────────────────────────────────────────────────────────

@dataclass
class _RawBlock:
    id: str
    duration_half_hours: int
    fte_requirement: int
    position_axes: dict[str, str]
    required_tool: ToolRequirement | None


def _parse_blocks(sheet, violations: list[Violation]) -> dict[str, _RawBlock]:
    raw: dict[str, _RawBlock] = {}
    if sheet is None:
        return raw

    rows = list(sheet.iter_rows())
    if not rows:
        violations.append(Violation(SHEET_BLOCKS, "-", "MISSING_HEADER_ROW",
                                    "first row is empty; expected required column headers"))
        return raw

    header_row = rows[0]
    header_index: dict[str, int] = {}
    for c, cell in enumerate(header_row):
        name = _read_string(cell)
        if name is None:
            continue
        key = name.strip().lower()
        header_index.setdefault(key, c)

    required_cols: dict[str, int] = {}
    missing_any = False
    for header in REQUIRED_BLOCKS_HEADERS:
        idx = header_index.get(header.lower())
        if idx is None:
            violations.append(Violation(SHEET_BLOCKS, "-", "BLOCKS_MISSING_HEADER",
                                        f"required column '{header}' is missing"))
            missing_any = True
        else:
            required_cols[header] = idx
    if missing_any:
        return raw

    id_col = required_cols[COL_HL_BLOCK]
    hrs_col = required_cols[COL_HRS]
    fte_col = required_cols[COL_FTE]
    tools_col = required_cols[COL_TOOLS]
    tools_amount_col = required_cols[COL_TOOLS_AMOUNT]

    seen_ids: set[str] = set()
    for r, row in enumerate(rows[1:], start=1):
        # Skip fully-empty rows silently
        if all(_read_string(row[c]) is None for c in required_cols.values() if c < len(row)):
            continue

        # ── HL Block ──
        id_str = _read_string(row[id_col]) if id_col < len(row) else None
        if id_str is None:
            violations.append(Violation(SHEET_BLOCKS, _cell_ref(r, id_col), "BLOCKS_MISSING_ID",
                                        f"row has data but '{COL_HL_BLOCK}' is empty"))
            continue
        block_id = id_str.strip()
        if block_id in seen_ids:
            violations.append(Violation(SHEET_BLOCKS, _cell_ref(r, id_col), "BLOCKS_DUPLICATE_ID",
                                        f"duplicate block id '{block_id}'"))
            continue
        seen_ids.add(block_id)

        # ── HRS ──
        hrs = _read_number(row[hrs_col]) if hrs_col < len(row) else None
        duration_hh: int = -1
        if hrs is None:
            violations.append(Violation(SHEET_BLOCKS, _cell_ref(r, hrs_col), "HRS_NOT_NUMERIC",
                                        "HRS must be a positive number; got empty/non-numeric"))
        elif hrs <= 0:
            violations.append(Violation(SHEET_BLOCKS, _cell_ref(r, hrs_col), "HRS_NOT_POSITIVE",
                                        f"HRS must be > 0; got {hrs}"))
        else:
            doubled = hrs * 2.0
            if abs(doubled - round(doubled)) > 1e-9:
                violations.append(Violation(SHEET_BLOCKS, _cell_ref(r, hrs_col), "HRS_NOT_HALF_HOUR",
                                            f"HRS must be a whole or half hour (×2 must be integer); got {hrs}"))
            else:
                duration_hh = int(round(doubled))

        # ── FTE ──
        fte_val = _read_number(row[fte_col]) if fte_col < len(row) else None
        fte_req: int = -1
        if fte_val is None:
            violations.append(Violation(SHEET_BLOCKS, _cell_ref(r, fte_col), "FTE_NOT_NUMERIC",
                                        "FTE must be a positive integer; got empty/non-numeric"))
        elif abs(fte_val - round(fte_val)) > 1e-9:
            violations.append(Violation(SHEET_BLOCKS, _cell_ref(r, fte_col), "FTE_NOT_INTEGER",
                                        f"FTE must be an integer; got {fte_val}"))
        elif fte_val < 1:
            violations.append(Violation(SHEET_BLOCKS, _cell_ref(r, fte_col), "FTE_NOT_POSITIVE",
                                        f"FTE must be ≥ 1; got {fte_val}"))
        else:
            fte_req = int(round(fte_val))

        # ── Tools ──
        tool_name = _read_string(row[tools_col]) if tools_col < len(row) else None
        tool_amount = _read_string(row[tools_amount_col]) if tools_amount_col < len(row) else None
        required_tool: ToolRequirement | None = None
        tool_present = tool_name is not None
        amount_present = tool_amount is not None
        if tool_present != amount_present:
            violations.append(Violation(SHEET_BLOCKS, _cell_ref(r, tools_col), "TOOLS_INCONSISTENT",
                                        "exactly one of TOOLS / TOOLS AMOUNT is empty"))
        elif tool_present:
            normalized = tool_name.strip().lower()  # type: ignore[union-attr]
            amount_norm = tool_amount.strip().lower()  # type: ignore[union-attr]
            if amount_norm == "one":
                required_tool = ToolRequirement(normalized, True)
            elif amount_norm == "multiple":
                required_tool = ToolRequirement(normalized, False)
            else:
                violations.append(Violation(SHEET_BLOCKS, _cell_ref(r, tools_amount_col),
                                            "TOOLS_AMOUNT_INVALID",
                                            f"TOOLS AMOUNT must be 'One' or 'Multiple'; got '{tool_amount}'"))

        # ── Position axes ──
        position_axes: dict[str, str] = {}
        for axis in POSITION_AXES:
            col = required_cols[axis]
            cell = row[col] if col < len(row) else None
            val = _read_string(cell)
            if val is None:
                continue
            if cell is not None and isinstance(cell.value, (int, float, bool)):
                violations.append(Violation(SHEET_BLOCKS, _cell_ref(r, col), "POSITION_NOT_STRING",
                                            f"position axis '{axis}' must be string-or-empty"))
                continue
            position_axes[axis] = val.strip().lower()

        if duration_hh > 0 and fte_req > 0:
            raw[block_id] = _RawBlock(
                id=block_id,
                duration_half_hours=duration_hh,
                fte_requirement=fte_req,
                position_axes=position_axes,
                required_tool=required_tool,
            )

    return raw


# ────────────────────────────────────────────────────────────────────────────
# TDM sheet (precedence)
# ────────────────────────────────────────────────────────────────────────────

def _parse_tdm(sheet, known_ids: set[str], violations: list[Violation]) -> dict[str, list[str]]:
    """Returns {block_id: [predecessor_ids]}."""
    predecessors: dict[str, list[str]] = defaultdict(list)
    if sheet is None:
        return predecessors
    rows = list(sheet.iter_rows())
    if not rows:
        violations.append(Violation(SHEET_TDM, "-", "MISSING_HEADER_ROW",
                                    "first row empty; expected mirrored block-name labels"))
        return predecessors

    header = rows[0]
    col_labels: list[str | None] = []
    for c in range(1, len(header)):
        s = _read_string(header[c])
        col_labels.append(s.strip() if s else None)

    row_labels: list[str | None] = []
    for r in range(1, len(rows)):
        s = _read_string(rows[r][0]) if rows[r] else None
        row_labels.append(s.strip() if s else None)

    if len(col_labels) != len(row_labels):
        violations.append(Violation(SHEET_TDM, "-", "TDM_NOT_SQUARE",
                                    f"row count {len(row_labels)} ≠ column count {len(col_labels)}"))
        return predecessors

    for i, (rl, cl) in enumerate(zip(row_labels, col_labels)):
        if rl != cl:
            violations.append(Violation(SHEET_TDM, _cell_ref(0, i + 1), "TDM_LABEL_MISMATCH",
                                        f"row label '{rl}' ≠ column label '{cl}' at index {i}"))

    for label in row_labels:
        if label is not None and label not in known_ids:
            violations.append(Violation(SHEET_TDM, "-", "TDM_UNKNOWN_BLOCK",
                                        f"label '{label}' not present in Blocks sheet"))

    # Cells: row r col c (both 1-based in matrix interior).
    # Convention: an X at (row r, col c) means col_labels[c-1] (the
    # predecessor) must finish before row_labels[r-1] (the successor).
    # Each row lists its predecessors in the columns.
    for r in range(1, len(rows)):
        row = rows[r]
        succ_label = row_labels[r - 1]
        if succ_label is None:
            continue
        for c in range(1, len(row)):
            if c - 1 >= len(col_labels):
                continue
            pred_label = col_labels[c - 1]
            if pred_label is None:
                continue
            if pred_label == succ_label:
                continue  # diagonal: silently ignore
            cell_val = _read_string(row[c])
            if cell_val is None:
                continue
            if cell_val.strip().lower() != "x":
                violations.append(Violation(SHEET_TDM, _cell_ref(r, c), "TDM_INVALID_MARKER",
                                            f"interior cell must be empty or 'X'; got '{cell_val}'"))
                continue
            if pred_label in known_ids and succ_label in known_ids:
                predecessors[succ_label].append(pred_label)

    # Cycle detection (Kahn)
    if not violations:
        in_deg: dict[str, int] = {bid: 0 for bid in known_ids}
        succ_map: dict[str, list[str]] = defaultdict(list)
        for succ_id, preds in predecessors.items():
            for p in preds:
                succ_map[p].append(succ_id)
                in_deg[succ_id] = in_deg.get(succ_id, 0) + 1
        queue: deque[str] = deque(b for b, d in in_deg.items() if d == 0)
        visited = 0
        while queue:
            n = queue.popleft()
            visited += 1
            for s in succ_map.get(n, ()):
                in_deg[s] -= 1
                if in_deg[s] == 0:
                    queue.append(s)
        if visited != len(known_ids):
            violations.append(Violation(SHEET_TDM, "-", "PRECEDENCE_CYCLE",
                                        "precedence graph contains a cycle"))

    return predecessors


# ────────────────────────────────────────────────────────────────────────────
# PDM sheet (spatial zones)
# ────────────────────────────────────────────────────────────────────────────

def _parse_pdm(sheet, known_ids: set[str], violations: list[Violation]) -> dict[str, set[str]]:
    """Returns {block_id: {zone_names}}."""
    zones_per_block: dict[str, set[str]] = defaultdict(set)
    if sheet is None:
        return zones_per_block
    rows = list(sheet.iter_rows())
    if not rows:
        violations.append(Violation(SHEET_PDM, "-", "MISSING_HEADER_ROW",
                                    "first row empty; expected zone-name column headers"))
        return zones_per_block

    header = rows[0]
    zone_labels: list[str | None] = []
    seen_zones: set[str] = set()
    for c in range(1, len(header)):
        s = _read_string(header[c])
        if s is None:
            zone_labels.append(None)
            continue
        trimmed = s.strip()
        if not trimmed:
            zone_labels.append(None)
            continue
        if trimmed in seen_zones:
            violations.append(Violation(SHEET_PDM, _cell_ref(0, c), "PDM_DUPLICATE_ZONE",
                                        f"duplicate zone '{trimmed}' (case-sensitive)"))
            zone_labels.append(None)
            continue
        seen_zones.add(trimmed)
        zone_labels.append(trimmed)

    for r in range(1, len(rows)):
        row = rows[r]
        block_label = _read_string(row[0]) if row else None
        if block_label is None:
            # Tolerate empty rows
            continue
        block_label = block_label.strip()
        if block_label not in known_ids:
            violations.append(Violation(SHEET_PDM, _cell_ref(r, 0), "PDM_UNKNOWN_BLOCK",
                                        f"label '{block_label}' not in Blocks sheet"))
            continue
        for c in range(1, len(row)):
            if c - 1 >= len(zone_labels):
                continue
            zone = zone_labels[c - 1]
            if zone is None:
                continue
            cell_val = _read_string(row[c])
            if cell_val is None:
                continue
            if cell_val.strip().lower() != "x":
                violations.append(Violation(SHEET_PDM, _cell_ref(r, c), "PDM_INVALID_MARKER",
                                            f"interior cell must be empty or 'X'; got '{cell_val}'"))
                continue
            zones_per_block[block_label].add(zone)

    return zones_per_block


# ────────────────────────────────────────────────────────────────────────────
# Public entry point
# ────────────────────────────────────────────────────────────────────────────

def load_instance(path: Any) -> list[Block]:
    """Load and validate one Excel instance. Raises :class:`LoaderError` on any violation."""
    p = Path(path)
    if not p.exists():
        raise LoaderError([Violation("-", "-", "FILE_UNREADABLE", f"file not found: {p}")])

    try:
        wb = load_workbook(filename=str(p), data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        raise LoaderError([Violation("-", "-", "FILE_UNREADABLE", str(e))]) from e

    violations: list[Violation] = []

    blocks_sheet = _find_sheet(wb, SHEET_BLOCKS)
    if blocks_sheet is None:
        violations.append(Violation("-", "-", "MISSING_SHEET",
                                    f"required sheet '{SHEET_BLOCKS}' not found"))
    tdm_sheet = _find_sheet(wb, SHEET_TDM)
    if tdm_sheet is None:
        violations.append(Violation("-", "-", "MISSING_SHEET",
                                    f"required sheet '{SHEET_TDM}' not found"))
    pdm_sheet = _find_sheet(wb, SHEET_PDM)
    if pdm_sheet is None:
        violations.append(Violation("-", "-", "MISSING_SHEET",
                                    f"required sheet '{SHEET_PDM}' not found"))

    if violations:
        raise LoaderError(violations)

    raw_blocks = _parse_blocks(blocks_sheet, violations)
    known_ids = set(raw_blocks.keys())
    predecessors = _parse_tdm(tdm_sheet, known_ids, violations)
    zones_per_block = _parse_pdm(pdm_sheet, known_ids, violations)

    if violations:
        raise LoaderError(violations)

    blocks: list[Block] = []
    for raw in raw_blocks.values():
        blocks.append(Block(
            id=raw.id,
            duration_half_hours=raw.duration_half_hours,
            fte_requirement=raw.fte_requirement,
            occupied_zones=frozenset(zones_per_block.get(raw.id, set())),
            position_axes=dict(raw.position_axes),
            required_tool=raw.required_tool,
            predecessor_block_ids=tuple(predecessors.get(raw.id, ())),
        ))
    return blocks
